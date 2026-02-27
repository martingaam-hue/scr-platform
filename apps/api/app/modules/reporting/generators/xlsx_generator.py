"""XLSX report generator using openpyxl."""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.modules.reporting.generators.base import BaseReportGenerator


class XLSXGenerator(BaseReportGenerator):
    """Generate Excel reports with cover sheet and data sections."""

    CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def generate(self, data: dict, sections: list[dict]) -> tuple[bytes, str]:
        wb = Workbook()
        self._create_cover_sheet(wb, data)

        for section in sections:
            self._create_section_sheet(wb, section, data)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), self.CONTENT_TYPE

    def _create_cover_sheet(self, wb: Workbook, data: dict) -> None:
        ws = wb.active
        ws.title = "Cover"

        # Brand color for header
        header_fill = PatternFill(
            start_color=self.brand_color.lstrip("#"),
            end_color=self.brand_color.lstrip("#"),
            fill_type="solid",
        )
        header_font = Font(color="FFFFFF", bold=True, size=16)

        ws.merge_cells("A1:E1")
        ws["A1"] = data.get("title", "Report")
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 40

        ws["A3"] = "Organization"
        ws["B3"] = self.org_name
        ws["A4"] = "Generated"
        ws["B4"] = self.generated_at
        ws["A5"] = "Report Type"
        ws["B5"] = self.template_config.get("audience", "general")

        if params := data.get("parameters"):
            ws["A7"] = "Parameters"
            ws["A7"].font = Font(bold=True)
            row = 8
            for key, val in params.items():
                if key != "output_format":
                    ws[f"A{row}"] = key.replace("_", " ").title()
                    ws[f"B{row}"] = str(val)
                    row += 1

        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def _create_section_sheet(self, wb: Workbook, section: dict, data: dict) -> None:
        name = section.get("name", "Section")
        # Excel sheet names max 31 chars
        sheet_name = name[:31]
        ws = wb.create_sheet(title=sheet_name)

        section_data = data.get(name, {})
        header_fill = PatternFill(
            start_color=self.brand_color.lstrip("#"),
            end_color=self.brand_color.lstrip("#"),
            fill_type="solid",
        )
        header_font = Font(color="FFFFFF", bold=True)

        if isinstance(section_data, list) and section_data:
            # Table data — list of dicts
            headers = list(section_data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row_idx, row_data in enumerate(section_data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

            # Auto-width
            for col_idx in range(1, len(headers) + 1):
                max_len = max(
                    len(str(ws.cell(row=r, column=col_idx).value or ""))
                    for r in range(1, len(section_data) + 2)
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        elif isinstance(section_data, dict):
            # Key-value pairs
            ws.cell(row=1, column=1, value="Metric").font = header_font
            ws["A1"].fill = header_fill
            ws.cell(row=1, column=2, value="Value").font = header_font
            ws["B1"].fill = header_fill

            for row_idx, (key, val) in enumerate(section_data.items(), 2):
                ws.cell(row=row_idx, column=1, value=key.replace("_", " ").title())
                ws.cell(row=row_idx, column=2, value=str(val) if val is not None else "")

            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 30

        else:
            # Plain text
            ws["A1"] = str(section_data) if section_data else f"No data for {name}"
